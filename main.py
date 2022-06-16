import calendar
import sys
import sqlalchemy
from datetime import datetime, date

from PyQt5 import QtWidgets, QtCore, QtGui
from PyQt5.QtCore import QStringListModel
from PyQt5.QtGui import QStandardItem, QStandardItemModel
from PyQt5.QtWidgets import QMessageBox
from sqlalchemy import text, and_, or_
from openpyxl import load_workbook

from choose import Ui_choose_dialog
from mainwindow import Ui_main_window
from workload import Ui_workload_dialog
from teacherworkload import Ui_teacher_workload
from teacheredit import Ui_teacher_edit
from directory import Ui_directory
from work_weekends_edit import Ui_work_weekends_form
from holydays_edit import Ui_holydays_form
from sicklist_edit import Ui_sicklist_edit
from vacation_edit import Ui_vacation_edit
from timetable import Ui_timetable_view
from replacement import Ui_Replacement_form
from data import db_session
from data.settings import Settings
from data.teachers import Teacher
from data.years import Years
from data.category import Category
from data.position import Position
from data.profession import Profession
from data.vacation_type import Vacation_type
from data.work_weekends import Work_weekend
from data.holydays import Holydays
from data.workload import Workload
from data.sick_list import Sick_list
from data.vacation import Vacation
from data.replacement import Replacement

months = {1: 'январь', 2: 'февраль', 3: 'март', 4: 'апрель', 5: 'май', 6: 'июнь',
          7: 'июль', 8: 'август', 9: 'сентябрь', 10: 'октябрь', 11: 'ноябрь', 12: 'декабрь'}

months_rp = {1: 'января', 2: 'февраля', 3: 'марта', 4: 'апреля', 5: 'мая', 6: 'июня',
             7: 'июля', 8: 'августа', 9: 'сентября', 10: 'октября', 11: 'ноября', 12: 'декабря'}


class VacationForm(QtWidgets.QDialog, Ui_vacation_edit):
    def __init__(self, db_sess):
        super().__init__()
        self.db_sess = db_sess
        self.ui = Ui_vacation_edit()
        self.ui.setupUi(self)
        self.model = None
        teachers = self.db_sess.query(Teacher).all()
        string_teachers = [str(t.id) + ":" + t.FIO for t in teachers]
        self.ui.teacher_cb.addItems(string_teachers)
        vacation_types = self.db_sess.query(Vacation_type).all()
        string_vtypes = [str(v.id) + ":" + v.short_name + " - " + v.full_name for v in vacation_types]
        self.ui.vacation_type.addItems(string_vtypes)
        self.set_controls(False)
        self.teacher_id = None
        self.vacation_type_id = None
        self.current_id = None
        self.current_action = None
        self.ui.add_button.clicked.connect(self.add_vacation)
        self.ui.edit_button.clicked.connect(self.edit_vacation)
        self.ui.save_button.clicked.connect(self.save_vacation)
        self.ui.del_button.clicked.connect(self.del_vacation)
        self.ui.teacher_cb.currentIndexChanged.connect(self.load_data)
        self.ui.vacation_type.currentIndexChanged.connect(self.load_data)
        self.ui.vacation_list.clicked.connect(self.table_change)
        self.load_data()

    def set_controls(self, mode):
        self.ui.date_begin.setEnabled(mode)
        self.ui.date_end.setEnabled(mode)
        self.ui.save_button.setEnabled(mode)
        self.ui.edit_button.setEnabled(not mode)
        self.ui.add_button.setEnabled(not mode)
        self.ui.del_button.setEnabled(not mode)

    def load_data(self):
        id, temp = self.ui.teacher_cb.currentText().split(':')
        self.teacher_id = int(id)
        id, temp = self.ui.vacation_type.currentText().split(':')
        self.vacation_type_id = int(id)
        vacations = self.db_sess.query(Vacation).filter(Vacation.teacher_id == self.teacher_id).all()
        vacations_str = [str(v.id) + ":" + v.date_begin.strftime("%d.%m.%Y") +
                         ' - ' + v.date_end.strftime("%d.%m.%Y") for v in vacations
                         if v.vacation_type == self.vacation_type_id]
        self.model = QStringListModel(vacations_str)
        self.ui.vacation_list.setModel(self.model)
        self.set_controls(False)

    def add_vacation(self):
        self.set_controls(True)
        self.current_action = 'add'
        self.ui.date_begin.setDate(datetime.now())
        self.ui.date_end.setDate(datetime.now())

    def edit_vacation(self):
        if self.current_id is not None:
            self.current_action = 'edit'
            self.set_controls(True)

    def table_change(self):
        data = self.model.data(self.ui.vacation_list.selectedIndexes()[0], 0)
        current_id, temp = data.split(':')
        self.current_id = int(current_id)
        date_begin, date_end = temp.split(' - ')
        self.ui.date_begin.setDate(datetime.strptime(date_begin, "%d.%m.%Y"))
        self.ui.date_end.setDate(datetime.strptime(date_end, "%d.%m.%Y"))

    def del_vacation(self):
        if self.current_id:
            valid = QMessageBox.question(
                self, '', "Действительно удалить элемент с № " + str(self.current_id),
                QMessageBox.Yes, QMessageBox.No)
            if valid == QMessageBox.Yes:
                v = self.db_sess.query(Vacation).filter(Vacation.id == self.current_id).one()
                self.db_sess.delete(v)
                self.db_sess.commit()
                self.load_data()

    def save_vacation(self):
        if self.current_action == 'add':
            v = Vacation(
                teacher_id=self.teacher_id,
                date_begin=self.ui.date_begin.date().toPyDate(),
                date_end=self.ui.date_end.date().toPyDate(),
                vacation_type=self.vacation_type_id
            )
            self.db_sess.add(v)
            self.db_sess.commit()
        elif self.current_action == 'edit':
            v = self.db_sess.query(Vacation).filter(Vacation.id == self.current_id).one()
            v.date_begin = self.ui.date_begin.date().toPyDate()
            v.date_end = self.ui.date_end.date().toPyDate()
            self.db_sess.commit()
        self.load_data()


class SickListForm(QtWidgets.QDialog, Ui_sicklist_edit):
    def __init__(self, db_sess):
        super().__init__()
        self.db_sess = db_sess
        self.ui = Ui_sicklist_edit()
        self.ui.setupUi(self)
        self.model = None
        teachers = self.db_sess.query(Teacher).all()
        string_teachers = [str(t.id) + ":" + t.FIO for t in teachers]
        self.ui.teacher_cb.addItems(string_teachers)
        self.set_controls(False)
        self.teacher_id = None
        self.current_id = None
        self.current_action = None
        self.ui.add_button.clicked.connect(self.add_sicklist)
        self.ui.edit_button.clicked.connect(self.edit_sicklist)
        self.ui.save_button.clicked.connect(self.save_sicklist)
        self.ui.del_button.clicked.connect(self.del_sicklist)
        self.ui.sicklist.clicked.connect(self.table_change)
        self.ui.teacher_cb.currentIndexChanged.connect(self.load_data)
        self.load_data()

    def set_controls(self, mode):
        self.ui.date_begin.setEnabled(mode)
        self.ui.date_end.setEnabled(mode)
        self.ui.save_button.setEnabled(mode)
        self.ui.edit_button.setEnabled(not mode)
        self.ui.add_button.setEnabled(not mode)
        self.ui.del_button.setEnabled(not mode)

    def load_data(self):
        id, temp = self.ui.teacher_cb.currentText().split(':')
        self.teacher_id = int(id)
        sicklist = self.db_sess.query(Sick_list).filter(Sick_list.teacher_id == self.teacher_id).all()
        sicklist_str = [str(sl.id) + ":" + sl.date_begin.strftime("%d.%m.%Y") + ' - ' +
                        sl.date_end.strftime("%d.%m.%Y")
                        for sl in sicklist]
        self.model = QStringListModel(sicklist_str)
        self.ui.sicklist.setModel(self.model)
        self.set_controls(False)

    def add_sicklist(self):
        self.set_controls(True)
        self.current_action = 'add'
        self.ui.date_begin.setDate(datetime.now())
        self.ui.date_end.setDate(datetime.now())

    def edit_sicklist(self):
        if self.current_id is not None:
            self.current_action = 'edit'
            self.set_controls(True)

    def table_change(self):
        data = self.model.data(self.ui.sicklist.selectedIndexes()[0], 0)
        current_id, temp = data.split(':')
        self.current_id = int(current_id)
        date_begin, date_end = temp.split(' - ')
        self.ui.date_begin.setDate(datetime.strptime(date_begin, "%d.%m.%Y"))
        self.ui.date_end.setDate(datetime.strptime(date_end, "%d.%m.%Y"))

    def del_sicklist(self):
        if self.current_id:
            valid = QMessageBox.question(
                self, '', "Действительно удалить элемент с № " + str(self.current_id),
                QMessageBox.Yes, QMessageBox.No)
            if valid == QMessageBox.Yes:
                sl = self.db_sess.query(Sick_list).filter(Sick_list.id == self.current_id).one()
                self.db_sess.delete(sl)
                self.db_sess.commit()
                self.load_data()

    def save_sicklist(self):
        if self.current_action == 'add':
            sl = Sick_list(
                teacher_id=self.teacher_id,
                date_begin=self.ui.date_begin.date().toPyDate(),
                date_end=self.ui.date_end.date().toPyDate()
            )
            self.db_sess.add(sl)
            self.db_sess.commit()
        elif self.current_action == 'edit':
            sl = self.db_sess.query(Sick_list).filter(Sick_list.id == self.current_id).one()
            sl.date_begin = self.ui.date_begin.date().toPyDate()
            sl.date_end = self.ui.date_end.date().toPyDate()
            self.db_sess.commit()
        self.load_data()


class WorkWeekendsForm(QtWidgets.QDialog, Ui_work_weekends_form):
    def __init__(self, db_sess):
        super().__init__()
        self.db_sess = db_sess
        self.ui = Ui_work_weekends_form()
        self.ui.setupUi(self)
        self.days_of_week = ['понедельник', 'вторник', 'среда', 'четверг', 'пятница']
        self.model = None
        for i, day in enumerate(self.days_of_week):
            self.ui.day_of_week.addItem(str(i) + ':' + day)
        self.load_data()
        self.ui.work_weekend.setDate(datetime.today())
        self.ui.add_button.clicked.connect(self.add_work_weekend)
        self.ui.edit_button.clicked.connect(self.edit_work_weekend)
        self.ui.save_button.clicked.connect(self.save_work_weekend)
        self.ui.work_days_list.clicked.connect(self.table_change)
        self.ui.del_button.clicked.connect(self.del_work_weekend)
        self.current_id = None
        self.current_action = None

    def load_data(self):
        work_weekends = self.db_sess.query(Work_weekend).order_by(Work_weekend.date_work).all()
        if len(work_weekends) > 0:
            string_work_weekends = [str(x.id) + ':' + x.date_work.strftime("%d.%m.%Y") + ' - ' +
                                    self.days_of_week[x.day_of_week] for x in work_weekends]
            self.model = QStringListModel(string_work_weekends)
            self.ui.work_days_list.setModel(self.model)

    def set_controls(self, mode):
        self.ui.work_weekend.setEnabled(mode)
        self.ui.day_of_week.setEnabled(mode)
        self.ui.save_button.setEnabled(mode)
        self.ui.edit_button.setEnabled(not mode)
        self.ui.add_button.setEnabled(not mode)
        self.ui.del_button.setEnabled(not mode)

    def add_work_weekend(self):
        self.set_controls(True)
        self.current_action = 'add'
        self.ui.work_weekend.setDate(datetime.now())

    def edit_work_weekend(self):
        if self.current_id is not None:
            self.current_action = 'edit'
            self.set_controls(True)

    def del_work_weekend(self):
        if self.current_id:
            valid = QMessageBox.question(
                self, '', "Действительно удалить элемент с № " + str(self.current_id),
                QMessageBox.Yes, QMessageBox.No)
            if valid == QMessageBox.Yes:
                ww = self.db_sess.query(Work_weekend).filter(Work_weekend.id == self.current_id)
                ww.delete()
                self.db_sess.commit()
                self.load_data()

    def save_work_weekend(self):
        if datetime.weekday(self.ui.work_weekend.date().toPyDate()) not in [5, 6]:
            msg = QMessageBox()
            msg.setIcon(QMessageBox.Critical)
            msg.setText("Введенная дата не является выходным днём")
            msg.setWindowTitle("Ошибка")
            msg.exec()
        else:
            if self.current_action == 'add':
                ww = Work_weekend(
                    date_work=self.ui.work_weekend.date().toPyDate(),
                    day_of_week=int(self.ui.day_of_week.currentText().split(':')[0])
                )
                self.db_sess.add(ww)
                self.db_sess.commit()
            elif self.current_action == 'edit':
                ww = self.db_sess.query(Work_weekend).filter(Work_weekend.id == self.current_id).one()
                ww.date_work = self.ui.work_weekend.date().toPyDate()
                ww.day_of_week = int(self.ui.day_of_week.currentText().split(':')[0])
                self.db_sess.commit()
            self.load_data()
        self.set_controls(False)
        self.current_action = None

    def table_change(self):
        data = self.model.data(self.ui.work_days_list.selectedIndexes()[0], 0)
        self.current_id, temp = data.split(':')
        self.current_id = int(self.current_id)
        work_weekend, day_of_week = temp.split(' - ')
        self.ui.work_weekend.setDate(datetime.strptime(work_weekend, "%d.%m.%Y"))
        self.ui.day_of_week.setCurrentIndex(self.days_of_week.index(day_of_week))


class HolydaysEditForm(QtWidgets.QDialog, Ui_holydays_form):
    def __init__(self, db_sess):
        super().__init__()
        self.db_sess = db_sess
        self.ui = Ui_holydays_form()
        self.ui.setupUi(self)
        self.ui.add_button.clicked.connect(self.add_holyday)
        self.ui.edit_button.clicked.connect(self.edit_holyday)
        self.ui.save_button.clicked.connect(self.save_holyday)
        self.ui.holydays_list.clicked.connect(self.table_change)
        self.ui.del_button.clicked.connect(self.del_holyday)
        self.current_action = None
        self.load_data()
        self.set_controls(False)

    def load_data(self):
        holydays = self.db_sess.query(Holydays).order_by(Holydays.holyday).all()
        if len(holydays) > 0:
            string_holydays = [str(x.id) + ':' + x.holyday.strftime("%d.%m.%Y") for x in holydays]
            self.model = QStringListModel(string_holydays)
            self.ui.holydays_list.setModel(self.model)

    def set_controls(self, mode):
        self.ui.holyday.setEnabled(mode)
        self.ui.save_button.setEnabled(mode)
        self.ui.edit_button.setEnabled(not mode)
        self.ui.add_button.setEnabled(not mode)
        self.ui.del_button.setEnabled(not mode)

    def add_holyday(self):
        self.current_action = 'add'
        self.set_controls(True)
        self.ui.holyday.setDate(datetime.now())

    def save_holyday(self):
        if datetime.weekday(self.ui.holyday.date().toPyDate()) not in range(5):
            msg = QMessageBox()
            msg.setIcon(QMessageBox.Critical)
            msg.setText("Введенная дата не является рабочим днём")
            msg.setWindowTitle("Ошибка")
            msg.exec()
        else:
            if self.current_action == 'add':
                hd = Holydays(
                    holyday=self.ui.holyday.date().toPyDate(),
                )
                self.db_sess.add(hd)
                self.db_sess.commit()
            elif self.current_action == 'edit':
                hd = self.db_sess.query(Holydays).filter(Holydays.id == self.current_id).one()
                hd.holyday = self.ui.holyday.date().toPyDate()
                self.db_sess.commit()
            self.load_data()
        self.set_controls(False)
        self.current_action = None

    def edit_holyday(self):
        self.current_action = 'edit'
        self.set_controls(True)

    def del_holyday(self):
        if self.current_id:
            valid = QMessageBox.question(
                self, '', "Действительно удалить элемент с № " + str(self.current_id),
                QMessageBox.Yes, QMessageBox.No)
            if valid == QMessageBox.Yes:
                hd = self.db_sess.query(Holydays).filter(Holydays.id == self.current_id)
                hd.delete()
                self.db_sess.commit()
                self.load_data()

    def table_change(self):
        data = self.model.data(self.ui.holydays_list.selectedIndexes()[0], 0)
        self.current_id, holyday = data.split(':')
        self.ui.holyday.setDate(datetime.strptime(holyday, "%d.%m.%Y"))


class DirectoryForm(QtWidgets.QDialog, Ui_directory):
    def __init__(self, db_sess):
        super().__init__()
        self.db_sess = db_sess
        self.ui = Ui_directory()
        self.ui.setupUi(self)
        self.ui.category_rb.clicked.connect(self.select_category)
        self.ui.position_rb.clicked.connect(self.select_position)
        self.ui.profession_rb.clicked.connect(self.select_profession)
        self.ui.vacation_type_rb.clicked.connect(self.select_vacation_type)
        self.ui.add_button.clicked.connect(self.add_item)
        self.ui.edit_button.clicked.connect(self.edit_item)
        self.ui.save_button.clicked.connect(self.save_item)
        self.ui.content_table.clicked.connect(self.table_change)
        self.current_action = None
        self.current_directory = None
        self.model = None

    def add_item(self):
        self.current_action = 'add'
        self.set_edit_mode(True)
        self.ui.short_name_edit.setText("")
        self.ui.full_value_edit.setText("")

    def edit_item(self):
        self.current_action = 'edit'
        self.set_edit_mode(True)

    def save_item(self):
        if self.current_action == 'add':
            if self.current_directory == 'category':
                cat = Category(
                    full_name=self.ui.full_value_edit.text(),
                    short_name=self.ui.short_name_edit.text()
                )
                self.db_sess.add(cat)
                self.db_sess.commit()
                self.select_category()
            elif self.current_directory == 'position':
                pos = Position(
                    full_name=self.ui.full_value_edit.text(),
                    short_name=self.ui.short_name_edit.text()
                )
                self.db_sess.add(pos)
                self.db_sess.commit()
                self.select_position()
            elif self.current_directory == 'profession':
                prof = Profession(
                    full_name=self.ui.full_value_edit.text(),
                    short_name=self.ui.short_name_edit.text()
                )
                self.db_sess.add(prof)
                self.db_sess.commit()
                self.select_profession()
            elif self.current_directory == 'vacation_type':
                vt = Vacation_type(
                    full_name=self.ui.full_value_edit.text(),
                    short_name=self.ui.short_name_edit.text()
                )
                self.db_sess.add(vt)
                self.db_sess.commit()
                self.select_vacation_type()
        elif self.current_action == 'edit':
            if self.current_directory == 'category':
                cat = self.db_sess.query(Category).filter(Category.id == self.current_id).first()
                cat.full_name = self.ui.full_value_edit.text()
                cat.short_name = self.ui.short_name_edit.text()
                self.db_sess.commit()
                self.current_action = None
                self.ui.save_button.setEnabled(False)
                self.select_category()
            elif self.current_directory == 'position':
                pos = self.db_sess.query(Position).filter(Position.id == self.current_id).first()
                pos.full_name = self.ui.full_value_edit.text()
                pos.short_name = self.ui.short_name_edit.text()
                self.db_sess.commit()
                self.current_action = None
                self.ui.save_button.setEnabled(False)
                self.select_position()
            elif self.current_directory == 'profession':
                prof = self.db_sess.query(Profession).filter(Profession.id == self.current_id).first()
                prof.full_name = self.ui.full_value_edit.text()
                prof.short_name = self.ui.short_name_edit.text()
                self.db_sess.commit()
                self.current_action = None
                self.ui.save_button.setEnabled(False)
                self.select_profession()
            elif self.current_directory == 'vacation_type':
                vt = self.db_sess.query(Vacation_type).filter(Vacation_type.id == self.current_id).first()
                vt.full_name = self.ui.full_value_edit.text()
                vt.short_name = self.ui.short_name_edit.text()
                self.db_sess.commit()
                self.current_action = None
                self.ui.save_button.setEnabled(False)
                self.select_vacation_type()
        self.set_edit_mode(False)

    def select_category(self):
        if self.ui.category_rb.isChecked():
            self.current_directory = 'category'
            category = self.db_sess.query(Category).all()
            if len(category) > 0:
                string_category = [str(x.id) + ':' + x.short_name + ' - ' + x.full_name for x in category]
                self.model = QStringListModel(string_category)
                self.ui.content_table.setModel(self.model)

    def select_vacation_type(self):
        if self.ui.vacation_type_rb.isChecked():
            self.current_directory = 'vacation_type'
            vacation_type = self.db_sess.query(Vacation_type).all()
            if len(vacation_type) > 0:
                string_vacation_type = [str(x.id) + ':' + x.short_name + ' - ' + x.full_name for x in vacation_type]
                self.model = QStringListModel(string_vacation_type)
                self.ui.content_table.setModel(self.model)

    def set_edit_mode(self, flag):
        if flag:
            self.ui.short_name_edit.setEnabled(True)
            self.ui.full_value_edit.setEnabled(True)
            self.ui.save_button.setEnabled(True)
            self.ui.add_button.setEnabled(False)
            self.ui.edit_button.setEnabled(False)
            if self.current_directory == 'category':
                self.ui.position_rb.setEnabled(False)
                self.ui.profession_rb.setEnabled(False)
                self.vacation_type_rb.setEnabled(False)
            elif self.current_directory == 'position':
                self.ui.category_rb.setEnabled(False)
                self.ui.profession_rb.setEnabled(False)
                self.vacation_type_rb.setEnabled(False)
            elif self.current_directory == 'profession':
                self.ui.category_rb.setEnabled(False)
                self.ui.position_rb.setEnabled(False)
                self.vacation_type_rb.setEnabled(False)
            elif self.current_directory == 'vacation_type':
                self.ui.category_rb.setEnabled(False)
                self.ui.position_rb.setEnabled(False)
                self.ui.profession_rb.setEnabled(False)
        else:
            self.ui.short_name_edit.setEnabled(False)
            self.ui.full_value_edit.setEnabled(False)
            self.ui.save_button.setEnabled(False)
            self.ui.add_button.setEnabled(True)
            self.ui.edit_button.setEnabled(True)
            self.ui.category_rb.setEnabled(True)
            self.ui.position_rb.setEnabled(True)
            self.ui.profession_rb.setEnabled(True)

    def select_position(self):
        if self.ui.position_rb.isChecked():
            self.current_directory = 'position'
            position = self.db_sess.query(Position).all()
            if len(position) > 0:
                string_position = [str(x.id) + ':' + x.short_name + ' - ' + x.full_name for x in position]
                self.model = QStringListModel(string_position)
                self.ui.content_table.setModel(self.model)

    def select_profession(self):
        if self.ui.profession_rb.isChecked():
            self.current_directory = 'profession'
            profession = self.db_sess.query(Profession).all()
            if len(profession) > 0:
                string_profession = [str(x.id) + ':' + x.short_name + ' - ' + x.full_name for x in profession]
                self.model = QStringListModel(string_profession)
                self.ui.content_table.setModel(self.model)
            else:
                self.model = QStringListModel([])
                self.ui.content_table.setModel(self.model)

    def table_change(self):
        data = self.model.data(self.ui.content_table.selectedIndexes()[0], 0)
        self.current_id, temp = data.split(':')
        self.current_id = int(self.current_id)
        short_name, full_name = temp.split(' - ')
        self.ui.short_name_edit.setText(short_name)
        self.ui.full_value_edit.setText(full_name)


class ReplacementEditForm(QtWidgets.QDialog, Ui_Replacement_form):
    def __init__(self, db_sess, year_id, month):
        super().__init__()
        self.ui = Ui_Replacement_form()
        self.ui.setupUi(self)
        self.db_sess = db_sess
        self.teachers = None
        self.replacement = None
        self.current_action = None
        self.current_id = None
        self.year = None
        self.substitute_id = None
        self.replaced_id = None
        self.model = None
        self.year_id = year_id
        self.month = month
        self.teachers = self.db_sess.query(Teacher).all()
        self.string_teachers = [str(t.id) + ":" + t.FIO for t in self.teachers]
        self.ui.substitute.addItems(self.string_teachers)
        self.ui.replaced.addItems(self.string_teachers)
        self.ui.add_button.clicked.connect(self.add_replacement)
        self.ui.edit_button.clicked.connect(self.edit_replacement)
        self.ui.save_button.clicked.connect(self.save_replacement)
        self.ui.del_button.clicked.connect(self.del_replacement)
        self.ui.replacements.clicked.connect(self.table_change)
        self.load_data()


    def load_data(self):
        years = self.db_sess.query(Years).filter(Years.id == self.year_id).one()
        if self.month > 8:
            self.year = years.begin_year
        else:
            self.year = years.end_year
        self.days_in_month = calendar.monthrange(self.year, self.month)[1]
        begin_month = date(self.year, self.month, 1)
        end_month = date(self.year, self.month, self.days_in_month)
        self.replacement = self.db_sess.query(Replacement).filter(and_(Replacement.date_repl >= begin_month,
                                                                       Replacement.date_repl <= end_month)).all()
        if len(self.replacement) > 0:
            self.model = QStandardItemModel()
            for r in self.replacement:
                row = [QStandardItem(str(r.id)), QStandardItem(self.teacher_by_id(r.substitute_id).FIO),
                       QStandardItem(self.teacher_by_id(r.replaced_id).FIO), QStandardItem(r.date_repl.strftime("%d.%m.%Y")),
                       QStandardItem(str(r.hours))]
                self.model.appendRow(row)
            self.ui.replacements.setModel(self.model)
            self.ui.replacements.verticalHeader().hide()
            self.model.setHorizontalHeaderItem(0, QStandardItem("№"))
            self.model.setHorizontalHeaderItem(1, QStandardItem("Заменяющий"))
            self.model.setHorizontalHeaderItem(2, QStandardItem("Заменяемый"))
            self.model.setHorizontalHeaderItem(3, QStandardItem("Дата"))
            self.model.setHorizontalHeaderItem(4, QStandardItem("Часы"))
            self.ui.replacements.setStyleSheet("QHeaderView::section { background-color:LightGray }")
            for i in range(5):
                self.ui.replacements.resizeColumnToContents(i)
        else:
            if self.model is not None:
                self.model.clear()
        self.set_controls(False)

    def table_change(self):
        index = self.ui.replacements.selectionModel().currentIndex()
        self.current_id = int(index.sibling(index.row(), 0).data())
        substitute = index.sibling(index.row(), 1).data()
        substitute = str(self.id_by_FIO(substitute).id) + ':' + substitute
        replaced = index.sibling(index.row(), 2).data()
        replaced = str(self.id_by_FIO(replaced).id) + ':' + replaced
        self.ui.substitute.setCurrentIndex(self.ui.substitute.findText(substitute))
        self.ui.replaced.setCurrentIndex(self.ui.replaced.findText(replaced))
        date_repl = index.sibling(index.row(), 3).data()
        hours = index.sibling(index.row(), 4).data()
        self.ui.date_repl.setDate(datetime.strptime(date_repl, '%d.%m.%Y'))
        self.ui.hours.setValue(int(hours))

    def teacher_by_id(self, id):
        for t in self.teachers:
            if t.id == id:
                return t
        return None

    def id_by_FIO(self, FIO):
        for t in self.teachers:
            if t.FIO == FIO:
                return t
        return None

    def set_controls(self, mode):
        self.ui.date_repl.setEnabled(mode)
        self.ui.save_button.setEnabled(mode)
        self.ui.substitute.setEnabled(mode)
        self.ui.replaced.setEnabled(mode)
        self.ui.hours.setEnabled(mode)
        self.ui.edit_button.setEnabled(not mode)
        self.ui.add_button.setEnabled(not mode)
        self.ui.del_button.setEnabled(not mode)

    def add_replacement(self):
        self.set_controls(True)
        self.ui.date_repl.setDate(datetime.now())
        self.current_action = 'Add'

    def edit_replacement(self):
        if self.current_id is not None:
            self.set_controls(True)
            self.current_action = 'Edit'

    def save_replacement(self):
        self.set_controls(False)
        if self.current_action == 'Add':
            self.substitute_id, _ = self.ui.substitute.currentText().split(':')
            self.substitute_id = int(self.substitute_id)
            self.replaced_id, _ = self.ui.replaced.currentText().split(':')
            self.replaced_id = int(self.replaced_id)
            replacement = Replacement(
                substitute_id=self.substitute_id,
                replaced_id=self.replaced_id,
                date_repl=self.ui.date_repl.date().toPyDate(),
                hours=self.ui.hours.value()
            )
            self.db_sess.add(replacement)
            self.db_sess.commit()
            self.load_data()
        elif self.current_action == 'Edit':
            replacement = self.db_sess.query(Replacement).filter(Replacement.id == self.current_id).first()
            id, _ = self.ui.substitute.currentText().split(':')
            replacement.substitute_id = int(id)
            id, _ = self.ui.replaced.currentText().split(':')
            replacement.replaced_id = int(id)
            replacement.date_repl = self.ui.date_repl.date().toPyDate()
            replacement.hours = self.ui.hours.value()
            self.db_sess.commit()
        self.current_action = None
        self.load_data()

    def del_replacement(self):
        if self.current_id:
            valid = QMessageBox.question(
                self, '', "Действительно удалить элемент с № " + str(self.current_id),
                QMessageBox.Yes, QMessageBox.No)
            if valid == QMessageBox.Yes:
                r = self.db_sess.query(Replacement).filter(Replacement.id == self.current_id)
                r.delete()
                self.db_sess.commit()
                self.current_id = None
                self.load_data()


class TeacherEditForm(QtWidgets.QDialog, Ui_teacher_edit):
    def __init__(self, db_sess):
        super().__init__()
        self.ui = Ui_teacher_edit()
        self.ui.setupUi(self)
        self.db_sess = db_sess
        self.teachers = self.db_sess.query(Teacher).all()
        self.category = self.db_sess.query(Category).all()
        self.position = self.db_sess.query(Position).all()
        self.profession = self.db_sess.query(Profession).all()
        self.current_action = None
        self.current_id = None
        if len(self.teachers) > 0:
            string_teacher = [str(x.id) + ':' + x.FIO for x in self.teachers]
            self.model = QStringListModel(string_teacher)
            self.ui.teacher_list.setModel(self.model)
        if len(self.category) > 0:
            string_category = [str(x.id) + ':' + x.short_name + '-' + x.full_name for x in self.category]
            self.ui.category_cb.clear()
            self.ui.category_cb.addItems(string_category)
        if len(self.position) > 0:
            string_position = [str(x.id) + ':' + x.short_name + '-' + x.full_name for x in self.position]
            self.ui.position_cb.clear()
            self.ui.position_cb.addItems(string_position)
        if len(self.profession) > 0:
            string_profession = [str(x.id) + ':' + x.short_name + '-' + x.full_name for x in self.profession]
            self.ui.profession_cb.clear()
            self.ui.profession_cb.addItems(string_profession)
        self.ui.add_button.clicked.connect(self.add_teacher)
        self.ui.edit_button.clicked.connect(self.edit_teacher)
        self.ui.save_button.clicked.connect(self.save_teacher)
        self.ui.teacher_list.clicked.connect(self.table_change)

    def set_controls(self, mode):
        self.ui.FIO.setEnabled(mode)
        self.ui.category_cb.setEnabled(mode)
        self.ui.position_cb.setEnabled(mode)
        self.ui.profession_cb.setEnabled(mode)
        self.ui.save_button.setEnabled(mode)
        self.ui.edit_button.setEnabled(not mode)
        self.ui.add_button.setEnabled(not mode)
        self.ui.del_button.setEnabled(not mode)

    def add_teacher(self):
        self.set_controls(True)
        self.ui.FIO.setText('')
        self.current_action = 'Add'

    def edit_teacher(self):
        self.set_controls(True)
        self.current_action = 'Edit'

    def refresh_list(self):
        self.teachers = self.db_sess.query(Teacher).all()
        if len(self.teachers) > 0:
            string_teacher = [str(x.id) + ':' + x.FIO for x in self.teachers]
            self.model = QStringListModel(string_teacher)
            self.ui.teacher_list.setModel(self.model)

    def save_teacher(self):
        self.set_controls(False)
        if self.current_action == 'Add':
            teacher = Teacher(
                FIO=self.ui.FIO.text(),
                position_id=int(self.ui.position_cb.currentText().split(':')[0]),
                category_id=int(self.ui.category_cb.currentText().split(':')[0]),
                profession_id=int(self.ui.profession_cb.currentText().split(':')[0])
            )
            self.db_sess.add(teacher)
            self.db_sess.commit()
            self.refresh_list()
        elif self.current_action == 'Edit':
            teacher = self.db_sess.query(Teacher).filter(Teacher.id == self.current_id).first()
            teacher.FIO = self.ui.FIO.text()
            teacher.position_id = int(self.ui.position_cb.currentText().split(':')[0])
            teacher.category_id = int(self.ui.category_cb.currentText().split(':')[0])
            teacher.profession_id = int(self.ui.profession_cb.currentText().split(':')[0])
            self.db_sess.commit()
        self.current_action = None
        self.refresh_list()

    def table_change(self):
        data = self.model.data(self.ui.teacher_list.selectedIndexes()[0], 0)
        self.current_id, temp = data.split(':')
        self.current_id = int(self.current_id)
        for t in self.teachers:
            if t.id == self.current_id:
                self.ui.FIO.setText(t.FIO)
                i = 0
                for pos in self.position:
                    if pos.id == t.position_id:
                        index = i
                        break
                    i += 1
                self.ui.position_cb.setCurrentIndex(index)
                i = 0
                for cat in self.category:
                    if cat.id == t.category_id:
                        index = i
                        break
                    i += 1
                self.ui.category_cb.setCurrentIndex(index)
                i = 0
                for prof in self.profession:
                    if prof.id == t.profession_id:
                        index = i
                        break
                    i += 1
                self.ui.profession_cb.setCurrentIndex(index)


class TeacherWorkLoadForm(QtWidgets.QDialog, Ui_teacher_workload):
    def __init__(self, teacher, workload):
        super().__init__()
        self.ui = Ui_teacher_workload()
        self.workload = workload
        self.ui.setupUi(self)
        self.ui.total.setText(str(sum(self.workload)))
        self.ui.teacher_name.setText(teacher)
        self.ui.d1.setValue(workload[0])
        self.ui.d2.setValue(workload[1])
        self.ui.d3.setValue(workload[2])
        self.ui.d4.setValue(workload[3])
        self.ui.d5.setValue(workload[4])
        self.ui.d1.valueChanged.connect(self.save_workload)
        self.ui.d2.valueChanged.connect(self.save_workload)
        self.ui.d3.valueChanged.connect(self.save_workload)
        self.ui.d4.valueChanged.connect(self.save_workload)
        self.ui.d5.valueChanged.connect(self.save_workload)

    def save_workload(self):
        self.workload[int(self.sender().objectName()[1:]) - 1] = self.sender().value()
        self.ui.total.setText(str(sum(self.workload)))


class TimetableForm(QtWidgets.QDialog, Ui_timetable_view):
    def __init__(self, db_sess, year_id, month):
        super().__init__()
        self.db_sess = db_sess
        self.year_id = year_id
        self.month = month
        self.ui = Ui_timetable_view()
        self.ui.setupUi(self)
        self.category = None
        self.position = None
        self.profession = None
        self.vacation_type = None
        self.load_data()
        self.model = None
        self.year = None
        self.days_in_month = None
        self.teachers = None
        self.workload = None
        self.work_weekend = None
        self.ui.out_to_excel.clicked.connect(self.export)
        self.load_data()

    def teacher_by_id(self, id):
        for t in self.teachers:
            if t.id == id:
                return t
        return None

    def teacher_by_name(self, name):
        for t in self.teachers:
            if t.FIO == name:
                return t

    def workload_by_id(self, id):
        result = []
        for w in self.workload:
            if w.teacher_id == id:
                result.append(w)
        return result

    def category_by_id(self, id):
        result = ""
        t = self.teacher_by_id(id)
        for c in self.category:
            if c.id == t.category_id:
                result = c.short_name
        return result

    def position_by_id(self, id):
        result = ""
        t = self.teacher_by_id(id)
        for p in self.position:
            if p.id == t.position_id:
                result = p.short_name
        return result

    def profession_by_id(self, id):
        result = ""
        t = self.teacher_by_id(id)
        for p in self.profession:
            if p.id == t.profession_id:
                result = p.short_name
        return result

    def vacation_type_by_id(self, id):
        result = ""
        for v in self.vacation_type:
            if v.id == id:
                result = v.short_name
        return result

    def load_data(self):
        years = self.db_sess.query(Years).filter(Years.id == self.year_id).one()
        if self.month > 8:
            self.year = years.begin_year
        else:
            self.year = years.end_year
        self.setWindowTitle('Просмотр табеля за ' + months[self.month] + ' ' + str(self.year))
        self.teachers = self.db_sess.query(Teacher).all()
        self.workload = self.db_sess.query(Workload).filter(Workload.year_id == self.year_id)
        self.category = self.db_sess.query(Category).all()
        self.position = self.db_sess.query(Position).all()
        self.profession = self.db_sess.query(Profession).all()
        self.vacation_type = self.db_sess.query(Vacation_type).all()
        w_end = self.db_sess.query(Work_weekend).all()
        self.work_weekend = dict()
        for ww in w_end:
            self.work_weekend[ww.date_work] = ww.day_of_week
        self.days_in_month = calendar.monthrange(self.year, self.month)[1]
        self.generate_new_model()
        self.ui.timetable.setModel(self.model)
        self.ui.timetable.verticalHeader().hide()
        self.ui.timetable.setStyleSheet("QHeaderView::section { background-color:LightGray }")
        self.model.setHorizontalHeaderItem(0, QStandardItem("ФИО"))
        for i in range(15):
            self.model.setHorizontalHeaderItem(i + 1, QStandardItem(str(i + 1)))
        self.model.setHorizontalHeaderItem(16, QStandardItem("1-15"))
        for i in range(17, self.days_in_month + 2):
            self.model.setHorizontalHeaderItem(i, QStandardItem(str(i - 1)))
        for i in range(self.days_in_month + 3):
            self.ui.timetable.resizeColumnToContents(i)
        self.model.setHorizontalHeaderItem(self.days_in_month + 2, QStandardItem("Всего"))

    def generate_new_model(self):
        begin_month = date(self.year, self.month, 1)
        end_month = date(self.year, self.month, self.days_in_month)
        begin_month_weekday = begin_month.weekday()
        self.model = QStandardItemModel()
        row = -1
        h_days = self.db_sess.query(Holydays).all()
        holydays = []
        for h in h_days:
            holydays.append(h.holyday)
        for t in self.teachers:
            workload = self.workload_by_id(t.id)
            sicklist = self.db_sess.query(Sick_list).filter((Sick_list.teacher_id == t.id) and
                                                            (((Sick_list.date_begin >= begin_month) and
                                                              (Sick_list.date_begin <= end_month)) or
                                                             ((Sick_list.date_end >= begin_month) and
                                                              (Sick_list.date_end <= end_month)))).all()
            vacations = self.db_sess.query(Vacation).filter(and_(Vacation.teacher_id == t.id,
                                                                 or_(and_(Vacation.date_begin >= begin_month,
                                                                          Vacation.date_begin <= end_month),
                                                                     and_(Vacation.date_end >= begin_month,
                                                                          Vacation.date_end <= end_month),
                                                                     and_(Vacation.date_begin <= begin_month,
                                                                          Vacation.date_end >= begin_month),
                                                                     and_(Vacation.date_begin <= end_month,
                                                                          Vacation.date_end >= end_month)
                                                                     ))).all()
            if len(workload) > 0:
                teacher_week = ["0"] * 5
                teacher_week += ["в", "в"]
                for w in workload:
                    teacher_week[w.day_of_week] = w.hours
                row += 1
                current_day = begin_month_weekday
                data = []
                for i in range(self.days_in_month):
                    curr_date = date(self.year, self.month, i + 1)
                    if len(vacations) > 0:
                        for v in vacations:
                            if v.date_begin <= curr_date <= v.date_end:
                                data.append(self.vacation_type_by_id(v.vacation_type))
                                break
                        else:
                            if curr_date in holydays:
                                data.append("в")
                            else:
                                if curr_date in self.work_weekend:
                                    data.append(str(teacher_week[self.work_weekend[curr_date]]))
                                else:
                                    data.append(str(teacher_week[current_day]))
                    else:
                        if len(sicklist) > 0:
                            for sl in sicklist:
                                if sl.date_begin <= curr_date <= sl.date_end:
                                    data.append("Б")
                                    break
                            else:
                                if curr_date in holydays:
                                    data.append("в")
                                else:
                                    if curr_date in self.work_weekend:
                                        data.append(str(teacher_week[self.work_weekend[curr_date]]))
                                    else:
                                        data.append(str(teacher_week[current_day]))
                        else:
                            if curr_date in holydays:
                                data.append("в")
                            else:
                                if curr_date in self.work_weekend:
                                    data.append(str(teacher_week[self.work_weekend[curr_date]]))
                                else:
                                    data.append(str(teacher_week[current_day]))
                    current_day += 1
                    current_day %= 7
                row_with_totals = [""] * (self.days_in_month + 2)
                total_sum = 0
                total_day = 0
                for i in range(15):
                    if data[i].isdigit():
                        total_sum += int(data[i])
                        total_day += 1
                    row_with_totals[i] = data[i]
                row_with_totals[15] = str(total_day) + "/" + str(total_sum)
                for i in range(15, len(data)):
                    if data[i].isdigit():
                        total_sum += int(data[i])
                        total_day += 1
                    row_with_totals[i + 1] = data[i]
                row_with_totals[len(data) + 1] = str(total_day) + "/" + str(total_sum)
                row_with_totals = [t.FIO] + row_with_totals
                items = []
                for item in row_with_totals:
                    st_it = QStandardItem(item)
                    if item == 'в':
                        st_it.setBackground(QtGui.QBrush(QtGui.QColor('pink')))
                    elif item == 'Б':
                        st_it.setBackground(QtGui.QBrush(QtGui.QColor('lightblue')))
                    elif item == 'О' or item == 'А':
                        st_it.setBackground(QtGui.QBrush(QtGui.QColor('lightgreen')))
                    elif '/' in item:
                        st_it.setBackground(QtGui.QBrush(QtGui.QColor('yellow')))
                    items.append(st_it)
                self.model.appendRow(items)

    def export(self):
        dates = ['AH', 'AL', 'AP', 'AT', 'AX', 'BB', 'BF', 'BJ', 'BN', 'BR', 'BV', 'BZ', 'CD', 'CH', 'CL',
                 'CW', 'DA', 'DE', 'DI', 'DM', 'DQ', 'DU', 'DY', 'EC', 'EG', 'EK', 'EO', 'ES', 'EW', 'FA', 'FE']
        wb = load_workbook('template22.xlsx')
        for ws in wb.worksheets:
            if ws.title == 'лист1':
                ws['BR4'] = self.days_in_month
                ws['BY4'] = months_rp[self.month]
                ws['CV4'] = str(self.year)[2:]
                for i in range(self.model.rowCount()):
                    for j in range(self.model.columnCount()):
                        if j == 0:
                            curr_teacher = self.teacher_by_name(self.model.item(i, j).text())
                            ws['A' + str(13 + i * 2)] = str(i + 1) + '. ' + curr_teacher.FIO
                            ws['M' + str(13 + i * 2)] = self.category_by_id(curr_teacher.id)
                            ws['Y' + str(13 + i * 2)] = self.position_by_id(curr_teacher.id)
                            ws['Y' + str(14 + i * 2)] = self.profession_by_id(curr_teacher.id)
                            ws['S' + str(13 + i * 2)] = sum([w.hours for w in self.workload_by_id(curr_teacher.id)])
                        elif j != 16 and j != self.model.columnCount() - 1:
                            if self.model.item(i, j).text()[0].isdigit():
                                cell_data = int(self.model.item(i, j).text())
                            else:
                                cell_data = self.model.item(i, j).text()
                            if j < 16:
                                ws[dates[j - 1] + str(14 + i * 2)] = cell_data
                            else:
                                ws[dates[j - 2] + str(14 + i * 2)] = cell_data
                        elif j == 16:
                            ws['CP' + str(14 + i * 2)] = self.model.item(i, j).text()
                        elif j == self.model.columnCount() - 1:
                            ws['FI' + str(14 + i * 2)] = self.model.item(i, j).text()
                file_name = 'табель' + datetime.strftime(datetime.now(), "%Y%m%d%H%M%S") + '.xlsx'
                wb.save(file_name)
                buttonReply = QMessageBox.information(self, "Информация",
                                                      'Табель сформирован в рабочем каталоге,\nимя файла "' +
                                                      file_name+'"', QMessageBox.Ok)
                break


class WorkLoadForm(QtWidgets.QDialog, Ui_workload_dialog):
    def __init__(self, db_sess, year_id):
        super().__init__()
        self.db_sess = db_sess
        self.ui = Ui_workload_dialog()
        self.ui.setupUi(self)
        self.ui.edit_workload.clicked.connect(self.edit_workload)
        self.ui.add_teacher.clicked.connect(self.add_teacher)
        self.ui.teachers_list.clicked.connect(self.table_change)
        self.year_id = year_id
        self.model = None
        self.current_teacher = None
        self.current_teacher_id = None
        self.workload = None
        self.load_data()

    def table_change(self):
        index = self.ui.teachers_list.selectionModel().currentIndex()
        self.current_teacher_id = int(index.sibling(index.row(), 0).data())
        self.current_teacher = index.sibling(index.row(), 1).data()
        self.workload = [0] * 5
        for i in range(5):
            self.workload[i] = int(index.sibling(index.row(), i + 2).data())

    def load_data(self):
        teachers = self.db_sess.query(Teacher).filter(text("id in (SELECT teacher_id from workload where year_id ="
                                                           + str(self.year_id) + ")")).all()
        if len(teachers) > 0:
            self.model = QStandardItemModel(len(teachers), 8)
            self.model.setHorizontalHeaderItem(0, QStandardItem("№"))
            self.model.setHorizontalHeaderItem(1, QStandardItem("ФИО"))
            self.model.setHorizontalHeaderItem(2, QStandardItem("Пн"))
            self.model.setHorizontalHeaderItem(3, QStandardItem("Вт"))
            self.model.setHorizontalHeaderItem(4, QStandardItem("Ср"))
            self.model.setHorizontalHeaderItem(5, QStandardItem("Чт"))
            self.model.setHorizontalHeaderItem(6, QStandardItem("Пт"))
            self.model.setHorizontalHeaderItem(7, QStandardItem("Всего"))
            row = 0
            for t in teachers:
                item = QStandardItem(str(t.id))
                self.model.setItem(row, 0, item)
                item = QStandardItem(str(t.FIO))
                self.model.setItem(row, 1, item)
                workload = ["0"] * 5
                hours = self.db_sess.query(Workload).filter(text("year_id=" + str(self.year_id) + " and teacher_id=" +
                                                                 str(t.id))).all()
                for h in hours:
                    workload[h.day_of_week] = str(h.hours)
                for i in range(5):
                    item = QStandardItem(str(workload[i]))
                    self.model.setItem(row, i + 2, item)
                s = sum([int(x) for x in workload])
                item = QStandardItem(str(s))
                self.model.setItem(row, 7, item)
                row += 1
        self.ui.teachers_list.setModel(self.model)
        self.ui.teachers_list.verticalHeader().hide()
        self.ui.teachers_list.setStyleSheet("QHeaderView::section { background-color:LightGray }")
        for i in range(8):
            self.ui.teachers_list.resizeColumnToContents(i)

    def edit_workload(self):
        if self.current_teacher:
            edit_workload_form = TeacherWorkLoadForm(self.current_teacher, self.workload)
            edit_workload_form.exec()
            if edit_workload_form.accepted:
                old_hours = self.db_sess.query(Workload).filter(Workload.teacher_id == self.current_teacher_id and
                                                                Workload.year_id == self.year_id).all()
                for i in range(5):
                    if self.workload[i] == 0:
                        for h in old_hours:
                            if h.day_of_week == i:
                                workload_id = h.id
                                wl = self.db_sess.query(Workload).filter(Workload.id == workload_id)
                                wl.delete()
                    elif self.workload[i] > 0:
                        for h in old_hours:
                            if h.day_of_week == i:
                                if h.hours != self.workload[i]:
                                    workload_id = h.id
                                    wl = self.db_sess.query(Workload).filter(Workload.id == workload_id).one()
                                    wl.hours = self.workload[i]
                                break
                        else:
                            wl = Workload(
                                year_id=self.year_id,
                                teacher_id=self.current_teacher_id,
                                day_of_week=i,
                                hours=self.workload[i]
                            )
                            self.db_sess.add(wl)
                self.db_sess.commit()
                self.load_data()

    def add_teacher(self):
        add_teacher_form = ChooseTeacherForm(self.db_sess, self.year_id)
        add_teacher_form.exec()
        if add_teacher_form.accepted:
            workload = [0] * 5
            edit_teacher_workload_form = TeacherWorkLoadForm(add_teacher_form.ui.choose_cb.currentText(), workload)
            edit_teacher_workload_form.exec()
            fio = edit_teacher_workload_form.ui.teacher_name.text()
            teacher_id = self.db_sess.query(Teacher).filter(Teacher.FIO == fio).one().id
            if edit_teacher_workload_form.accepted:
                for i in range(5):
                    if edit_teacher_workload_form.workload[i] > 0:
                        wl = Workload(
                            year_id=self.year_id,
                            teacher_id=teacher_id,
                            day_of_week=i,
                            hours=edit_teacher_workload_form.workload[i]
                        )
                        self.db_sess.add(wl)
                self.db_sess.commit()
                self.load_data()


class ChooseYearForm(QtWidgets.QDialog, Ui_choose_dialog):
    def __init__(self):
        super().__init__()
        self.ui = Ui_choose_dialog()
        self.ui.setupUi(self)
        self.setWindowTitle('Выбор года')
        self.ui.choose_label.setText('Выберите год')
        for i in range(2021, 2041):
            self.ui.choose_cb.addItem(str(i) + '-' + str(i + 1))


class ChooseMonthForm(QtWidgets.QDialog, Ui_choose_dialog):
    def __init__(self):
        super().__init__()
        self.ui = Ui_choose_dialog()
        self.ui.setupUi(self)
        self.setWindowTitle('Выбор месяца')
        self.ui.choose_label.setText('Выберите месяц')
        for i in range(1, 13):
            self.ui.choose_cb.addItem(months[i])


class ChooseTeacherForm(QtWidgets.QDialog, Ui_choose_dialog):
    def __init__(self, db_sess, year_id):
        self.db_sess = db_sess
        super().__init__()
        self.ui = Ui_choose_dialog()
        self.ui.setupUi(self)
        self.setWindowTitle('Выбор учителя')
        self.ui.choose_label.setText('Выберите учителя')
        teachers = self.db_sess.query(Teacher).filter(text("id not in (SELECT teacher_id from workload where year_id ="
                                                           + str(year_id) + ")")).all()
        if len(teachers) > 0:
            for t in teachers:
                self.ui.choose_cb.addItem(t.FIO)


class MainWindow(QtWidgets.QDialog, Ui_main_window):
    def __init__(self):
        super().__init__()
        self.ui = Ui_main_window()
        db_session.global_init("timetable.db")
        self.ui.setupUi(self)
        self.ui.choose_year_button.clicked.connect(self.choose_year)
        self.ui.choose_month_button.clicked.connect(self.choose_month)
        self.ui.workload_button.clicked.connect(self.workload_edit)
        self.ui.directories_button.clicked.connect(self.directories_edit)
        self.ui.teachers_button.clicked.connect(self.teachers_edit)
        self.ui.work_weekends_button.clicked.connect(self.work_weekends_edit)
        self.ui.holydays_button.clicked.connect(self.holydays_edit)
        self.ui.sick_list_button.clicked.connect(self.sicklist_edit)
        self.ui.vacation_button.clicked.connect(self.vacation_edit)
        self.ui.replacement_button.clicked.connect(self.replacement_edit)
        self.ui.make_timetable_button.clicked.connect(self.timetable)
        self.db_sess = db_session.create_session()
        self.current_year_end = None
        self.current_year_begin = None
        self.current_year_id = None
        self.load_settings()

    def load_settings(self):
        settings = self.db_sess.query(Settings).all()
        # p = self.db_sess.query(Settings).count()
        sett_dict = dict()
        for row in settings:
            sett_dict[row.parameter] = row.value
        if 'year' in sett_dict:
            self.current_year_id = int(sett_dict['year'])
            curr_year = self.db_sess.query(Years).filter(Years.id == self.current_year_id).one()
            self.current_year_begin = curr_year.begin_year
            self.current_year_end = curr_year.end_year
            self.ui.year_label.setText('Текущий год ' + str(self.current_year_begin) + '-' +
                                       str(self.current_year_end))
        if 'month' in sett_dict:
            self.current_month = int(sett_dict['month'])
            self.ui.month_label.setText('Текущий месяц ' + months[self.current_month])

    def teachers_edit(self):
        teacher_edit_form = TeacherEditForm(self.db_sess)
        teacher_edit_form.exec()

    def holydays_edit(self):
        if not self.current_year_id:
            msg = QMessageBox()
            msg.setIcon(QMessageBox.Critical)
            msg.setText("Не выбран текущий год")
            msg.setWindowTitle("Ошибка")
            msg.exec()
        else:
            holydays_edit_form = HolydaysEditForm(self.db_sess)
            holydays_edit_form.exec()

    def sicklist_edit(self):
        sicklist_edit_form = SickListForm(self.db_sess)
        sicklist_edit_form.exec()

    def timetable(self):
        timetable_form = TimetableForm(self.db_sess, self.current_year_id, self.current_month)
        timetable_form.exec()

    def vacation_edit(self):
        vacation_form = VacationForm(self.db_sess)
        vacation_form.exec()

    def work_weekends_edit(self):
        work_weekends_edit_form = WorkWeekendsForm(self.db_sess)
        work_weekends_edit_form.exec()

    def choose_year(self):
        choose_form = ChooseYearForm()
        choose_form.exec()
        if choose_form.accepted:
            self.current_year_begin, self.current_year_end = map(int,
                                                                 choose_form.ui.choose_cb.currentText().split('-'))
            self.ui.year_label.setText('Текущий год ' + str(choose_form.ui.choose_cb.currentText()))
            try:
                curr_year = self.db_sess.query(Years).filter(Years.begin_year == self.current_year_begin
                                                             and Years.end_year == self.current_year_end).one()
            except sqlalchemy.exc.NoResultFound:
                year = Years(
                    begin_year=self.current_year_begin,
                    end_year=self.current_year_end
                )
                self.db_sess.add(year)
                self.db_sess.commit()
                curr_year = self.db_sess.query(Years).filter(Years.begin_year == self.current_year_begin
                                                             and Years.end_year == self.current_year_end).one()
            self.current_year_id = curr_year.id
            self.save_setting("year", str(self.current_year_id))

    def save_setting(self, parameter, value):
        try:
            setting = self.db_sess.query(Settings).filter(Settings.parameter == parameter).one()
            self.db_sess.query(Settings).filter(Settings.parameter == parameter).update({"value": value})
            self.db_sess.commit()
        except sqlalchemy.exc.NoResultFound:
            setting = Settings(
                parameter=parameter,
                value=value
            )
            self.db_sess.add(setting)
            self.db_sess.commit()

    def choose_month(self):
        choose_form = ChooseMonthForm()
        choose_form.exec()
        if choose_form.accepted:
            self.current_month_str = choose_form.ui.choose_cb.currentText()
            self.ui.month_label.setText('Текущий месяц ' + self.current_month_str)
            for key in months.keys():
                if months[key] == self.current_month_str:
                    self.current_month = key
            self.save_setting("month", self.current_month)

    def workload_edit(self):
        workload_edit_form = WorkLoadForm(self.db_sess, self.current_year_id)
        workload_edit_form.exec()

    def replacement_edit(self):
        replacement_edit_form = ReplacementEditForm(self.db_sess, self.current_year_id, self.current_month)
        replacement_edit_form.exec()

    def directories_edit(self):
        directories_edit_form = DirectoryForm(self.db_sess)
        directories_edit_form.exec()


def except_hook(cls, exception, traceback):
    sys.__excepthook__(cls, exception, traceback)


app = QtWidgets.QApplication([])
application = MainWindow()
application.show()
sys.excepthook = except_hook
sys.exit(app.exec())
